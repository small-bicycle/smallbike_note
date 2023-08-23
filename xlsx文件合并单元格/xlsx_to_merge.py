#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
@File  : xlsx_to_merge.py
@Author: smallbike
@Date  : 2023/8/23
@Desc  :  xlsx文件合并单元格
'''

import os
import openpyxl


def merge_xlsx(file):
    '''
    合并单元格
    :param file: 需要合并的文件
    :return:
    '''
    # 打开文件目录下的文件的第一张sheet页
    wb = openpyxl.load_workbook(os.path.join(path, file))
    sheet = wb["Sheet1"]
    print(sheet.max_row)  # 总行数
    print(sheet.max_column)  # 总列数
    # 将 所有的 列号保存下来
    column_list = []
    for i in range(1, sheet.max_column + 1):
        value = sheet.cell(row=2, column=i)
        # value.coordinate 获取当前值所在的坐标 ---> A1, B1, C1
        column_list.append(value.coordinate[0])
    print(column_list)
    # 遍历每一列 去合并
    for letter in column_list:
        sumList = []
        print(sumList)
        print(letter)
        # 先把每一列列的所有内容放入sumList中
        for i in range(2, sheet.max_row + 1):
            value = sheet.cell(row=i, column=1).value
            if value:
                sumList.append(value)
            else:
                sumList.append('')
        print(sumList)
        # 开始合并单元格
        prow = 0
        frow = 0
        flag = sumList[0]
        for i in range(len(sumList)):
            if sumList[i] != flag:
                flag = sumList[i]
                frow = i - 1
                if frow >= prow:
                    # merge_cells() ---> 合并指定坐标
                    sheet.merge_cells(f"{letter}{prow + 2}:{letter}{frow + 2}")
                    prow = frow + 1
            if i == len(sumList) - 1:
                frow = i
                sheet.merge_cells(f"{letter}{prow + 2}:{letter}{frow + 2}")

        print(f'合并完成一列： {letter}')
    # 保存表
    wb.save(os.path.join(path, file))
    print(f'操作完毕')


if __name__ == '__main__':
    # 需要操作的文件
    file_name = 'test_files.xlsx'
    # 获得当前py文件的路径
    path = os.getcwd()
    print(os.path.join(path, file_name))
    merge_xlsx(file=file_name)
