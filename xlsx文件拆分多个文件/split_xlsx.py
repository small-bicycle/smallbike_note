#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
@File  : split_xlsx.py
@Author: smallbike
@Email : 15797642529@163.com
@Date  : 2023/8/23 8:41
@Desc  :  把表格按照 指定行数 拆分成 多个文件
'''

import openpyxl


def xlsx_split(filtes_path, cols_limit, is_tite=None):
    """

    :param filtes_path: 目标文件
    :param cols_limit: 拆分行数的数量
    :param is_tite: 是否需要表头
    :return:
    """
    workbook = openpyxl.load_workbook(filename=filtes_path)
    sheet_origin = workbook.active
    m_rows = len([row for row in sheet_origin if not all([cell.value is None for cell in row])])
    m_cols = sheet_origin.max_column

    print(f'最大行数：【{m_rows}】 - 最大列数：【{m_cols}】')
    sheets = m_rows / cols_limit
    if not sheets.is_integer():
        sheets = int(sheets) + 1

    print(f' 分成的份数【{sheets}】')

    for i in range(1, sheets + 1):
        wb = openpyxl.Workbook()
        sheets = wb['Sheet']

        start_num = 1
        if is_tite:
            start_num = 2
            for n in range(1, m_cols + 1):
                sheets.cell(row=1, colume=n).value = sheet_origin.cell(row=1, colume=n).value

        t = start_num + cols_limit * (i - 1)
        num_index = start_num
        for row_num in range(t, t + cols_limit + start_num - 1):
            for col_num in range(1, m_cols + 1):
                sheets.cell(row=num_index, column=col_num).value = sheet_origin.cell(row=row_num, column=col_num).value
            num_index = num_index + 1
        wb.save(f"{filtes_path}_{i}.xlsx")


if __name__ == '__main__':
    cols_limit = 5000
    filtes_path = r'test_files.xlsx'
    xlsx_split(filtes_path, cols_limit)
